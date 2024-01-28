from openpyxl import Workbook
from Kpop_artist import *
import pymysql

def write_excel(file_name, sheet_title, arr_column, data):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet(sheet_title)

    # write_ws = write_ws.active
    write_ws.append(arr_column)
    
    for val in data:
        # info = val.write_arr()
        write_ws.append(val)
    write_wb.save("C:/study/snc/market_crolling/python/refer_weverse/{}.xlsx".format(file_name))

def write_goods_excel(sheet_title,data):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet(sheet_title)

    write_ws = write_wb.active
    write_ws.append(["id","artist_id", "title", "category", "price", "option_size", "option_ment", "notice", "main_img", "sub_img","date","option_max"])
    
    for val in data:
        # info = val.write_arr()
        write_ws.append(val)
    write_wb.save("C:/study/snc/market_crolling/python/refer_weverse/goods.xlsx")

def write_notice_excel(sheet_title,data):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet(sheet_title)

    write_ws = write_wb.active
    write_ws.append(["id", "artist_id", "category", "title", "content", "date"])
    
    for val in data:
        # info = val.write_arr()
        write_ws.append(val)
    write_wb.save("C:/study/snc/market_crolling/python/refer_weverse/notice.xlsx")

def db_connect():
    cart_db = pymysql.connect(
        user='root',
        password='@gksrudwnGOD74',
        host='127.0.0.1',
        db='shopdb',
        charset='utf8'
    )
    return cart_db

def  write_sql(tuple_data, insert_sql):
    cart_db = db_connect()

    with cart_db:
        cur = cart_db.cursor()
        cur.execute(insert_sql,tuple_data)

        cart_db.commit()

def artist_write(index, data):
    with open("./python/refer_weverse/artist_sql.csv","w",encoding="utf-8") as file :
        file.write("name\tkorean_name\tmember\tmv\tartist_img\tdebut\tsns")
        for index, val in data:
            file.write()